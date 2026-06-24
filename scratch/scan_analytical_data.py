import os
import re
import string
import openpyxl
from openpyxl import load_workbook

# 1. Helper to extract readable text from binary .doc files
def extract_text_from_doc(file_path):
    try:
        with open(file_path, 'rb') as f:
            content = f.read()
        # Find sequences of printable ASCII characters of length >= 4
        # Since .doc stores text in UTF-16 or ASCII, we extract both
        ascii_strings = re.findall(rb'[a-zA-Z0-9\s\.,;:!\?\-\(\)\/\*\"\'\=\#\_]{4,}', content)
        decoded = []
        for s in ascii_strings[:1000]: # limit to avoid huge outputs
            try:
                text = s.decode('ascii').strip()
                if len(text) > 6:
                    decoded.append(text)
            except Exception:
                pass
        
        # Check UTF-16
        utf16_content = content.decode('utf-16', errors='ignore')
        utf16_strings = re.findall(r'[a-zA-Z0-9\s\.,;:!\?\-\(\)\/\*\"\'\=\#\_]{4,}', utf16_content)
        for s in utf16_strings[:1000]:
            text = s.strip()
            if len(text) > 6 and text not in decoded:
                decoded.append(text)
                
        return "\n".join(decoded[:100]) # return first 100 lines
    except Exception as e:
        return f"Error extracting from {file_path}: {e}"

# 2. Helper to scan Excel structures
def scan_excel(file_path):
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        info = f"Workbook: {os.path.basename(file_path)}\nSheets: {', '.join(sheets)}\n"
        for name in sheets[:3]: # inspect first few sheets
            ws = wb[name]
            rows = []
            for r in ws.iter_rows(max_row=5, values_only=True):
                rows.append([str(c)[:25] if c is not None else '' for c in r])
            info += f"  Sheet: {name}\n"
            for r in rows:
                info += f"    {r}\n"
        return info
    except Exception as e:
        return f"Error reading {file_path}: {e}"

print("=== SCANNING ANNEXURE & SOPs ===")
sop_dir = r"c:\Users\acer\Documents\projects\RAIS-Pro\ANALYTICAL DATA\SOP"
for f in os.listdir(sop_dir):
    if f.endswith(".doc"):
        path = os.path.join(sop_dir, f)
        print(f"\n--- FILE: {f} ---")
        text = extract_text_from_doc(path)
        print(text[:1500])

print("\n=== SCANNING EXCEL WORKBOOKS ===")
rej_dir = r"c:\Users\acer\Documents\projects\RAIS-Pro\ANALYTICAL DATA\REJECTION ANALYSIS 2025-26"
yearly_path = os.path.join(rej_dir, "YEARLY ANALYSIS.xlsx")
if os.path.exists(yearly_path):
    print(scan_excel(yearly_path))

visual_dir = r"c:\Users\acer\Documents\projects\RAIS-Pro\ANALYTICAL DATA\SIZE WISE REJECTION\VISUAL"
comm_visual = os.path.join(visual_dir, "commulative 2026-27.xlsx")
if os.path.exists(comm_visual):
    print(scan_excel(comm_visual))

valve_dir = r"c:\Users\acer\Documents\projects\RAIS-Pro\ANALYTICAL DATA\SIZE WISE REJECTION\VALVE INTEGRITY"
comm_valve = os.path.join(valve_dir, "commulative 2026-27.xlsx")
if os.path.exists(comm_valve):
    print(scan_excel(comm_valve))
